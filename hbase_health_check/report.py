import datetime
import gzip
from hbase_health_check.models import HBaseCluster
from pathlib import Path
from dacite import from_dict
from openpyxl import load_workbook
import statistics
from hbase_health_check.configuration import *

import orjson


def bytes_to_human_readable(num_bytes):
    num_bytes = int(num_bytes)
    # Define the suffixes for each size scale
    suffixes = ['B', 'KB', 'MB', 'GB', 'TB', 'PB', 'EB', 'ZB', 'YB']

    # Calculate the logarithmic value to determine the size scale
    i = 0
    while num_bytes >= 1024 and i < len(suffixes) - 1:
        num_bytes /= 1024.
        i += 1

    # Format the number to two decimal places and append the appropriate suffix
    return f"{num_bytes:.2f} {suffixes[i]}"


def report_health_check():
    if gz_file_path.exists():
        clusters = gzip.open(gz_file_path, 'rb').read()
    elif txt_file_path.exists():
        clusters = open(txt_file_path, 'rb').read()
    else:
        print('Error no output file found.')
        return

    workbook = load_workbook('report-template.xlsx')
    clusters = orjson.loads(clusters)
    for _cluster in clusters:
        cluster = from_dict(data_class=HBaseCluster, data=_cluster)
        current_row = 3
        sheet = workbook.copy_worksheet(from_worksheet=workbook['template'])
        sheet.title = cluster.cluster_id

        # region count
        for region_server in cluster.region_servers:
            sheet.insert_rows(current_row)
            sheet.cell(row=current_row, column=2, value=region_server.instance_dns_name)
            sheet.cell(row=current_row, column=3, value=str(region_server.region_num))
            current_row += 1

        current_row += 2

        # region size
        for region_server in cluster.region_servers:
            for region in region_server.regions:
                if region.store_file_size_gb >= LARGE_REGION_SIZE:
                    sheet.insert_rows(current_row)
                    sheet.cell(row=current_row, column=2, value=region_server.instance_dns_name)
                    sheet.cell(row=current_row, column=3, value=f'{region.store_file_size_gb:2f}GB')
                    sheet.cell(row=current_row, column=4, value=region.table_name)
                    sheet.cell(row=current_row, column=5, value=region.name)
                    current_row += 1

        current_row += 2

        # region server request
        data = [region_server.total_requests for region_server in cluster.region_servers]
        max_req = max(data)
        min_req = min(data)
        mean_req = statistics.mean(data)
        stddev_req = statistics.stdev(data) if len(data) > 1 else 0
        is_first = True
        for region_server in cluster.region_servers:
            sheet.insert_rows(current_row)
            sheet.cell(row=current_row, column=2, value=region_server.instance_dns_name)
            sheet.cell(row=current_row, column=3, value=str(region_server.total_requests))
            if is_first:
                sheet.cell(row=current_row, column=4, value=str(max_req))
                sheet.cell(row=current_row, column=5, value=str(min_req))
                sheet.cell(row=current_row, column=6, value=str(mean_req))
                sheet.cell(row=current_row, column=7, value=str(stddev_req))
                is_first = False
            current_row += 1

        current_row += 2

        # column family count
        all_tables = {}
        for region_server in cluster.region_servers:
            for table in region_server.tables:
                cf = all_tables.get(table.name, 0)
                if table.column_families > COLUMN_FAMILY_THRESHOLD and table.column_families > cf:
                    all_tables[table.name] = table.column_families
        for table_name, cf in all_tables.items():
            sheet.insert_rows(current_row)
            sheet.cell(row=current_row, column=2, value=table_name)
            sheet.cell(row=current_row, column=3, value=str(cf))
            current_row += 1

        current_row += 2
        # skewed table
        all_tables = {}
        for idx, region_server in enumerate(cluster.region_servers):
            for table in region_server.tables:
                data = all_tables.get(table.name, [0]*len(cluster.region_servers))
                data[idx] = table.total_requests
                all_tables[table.name] = data
        # print(all_tables)
        for table_name, data in all_tables.items():
            min_req = min(data)
            max_req = max(data)
            mean_req = statistics.mean(data)
            stddev_req = statistics.stdev(data) if len(data) > 1 else 0
            if max_req > mean_req * SKEWNESS_THRESHOLD:
                sheet.insert_rows(current_row)
                sheet.cell(row=current_row, column=2, value=table_name)
                sheet.cell(row=current_row, column=3, value=max_req)
                sheet.cell(row=current_row, column=4, value=min_req)
                sheet.cell(row=current_row, column=5, value=mean_req)
                sheet.cell(row=current_row, column=6, value=stddev_req)
                current_row += 1

        current_row += 2
        config_map = {'HBASE_HEAPSIZE': None}
        for config in CONFIGURATIONS:
            config_map[config] = None

        for cluster_config in cluster.configurations:
            if cluster_config['Classification'] == 'hbase-env':
                for conf in cluster_config['Configurations']:
                    if 'HBASE_HEAPSIZE' in conf['Properties']:
                        config_map['HBASE_HEAPSIZE'] = conf['Properties']['HBASE_HEAPSIZE']
            elif cluster_config['Classification'] == 'hbase-site':
                for k, v in cluster_config['Properties'].items():
                    if k in CONFIGURATIONS:
                        if k.endswith('size'):
                            try:
                                v = bytes_to_human_readable(v)
                            except Exception as e:
                                pass
                        config_map[k] = v
        for k, v in config_map.items():
            sheet.insert_rows(current_row)
            sheet.cell(row=current_row, column=2, value=k)
            sheet.cell(row=current_row, column=3, value=v)
            sheet.cell(row=current_row, column=4, value=CONFIGURATIONS[k])
            current_row += 1

        current_row += 2
        for region_server in cluster.region_servers:
            sheet.insert_rows(current_row)
            sheet.cell(row=current_row, column=2, value=region_server.instance_dns_name)
            sheet.cell(row=current_row, column=3, value=str(region_server.store_files))
            current_row += 1

        current_row += 2
        for region_server in cluster.region_servers:
            sheet.insert_rows(current_row)
            sheet.cell(row=current_row, column=2, value=region_server.instance_dns_name)
            sheet.cell(row=current_row, column=3, value=str(region_server.l1_hit_ratio))
            sheet.cell(row=current_row, column=4, value=str(region_server.l2_hit_ratio))
            current_row += 1

        current_row += 2
        for region_server in cluster.region_servers:
            sheet.insert_rows(current_row)
            sheet.cell(row=current_row, column=2, value=region_server.instance_dns_name)
            sheet.cell(row=current_row, column=3, value=str(region_server.split_num))
            current_row += 1

        current_row += 2
        for region_server in cluster.region_servers:
            sheet.insert_rows(current_row)
            sheet.cell(row=current_row, column=2, value=region_server.instance_dns_name)
            sheet.cell(row=current_row, column=3, value=str(region_server.gc_num))
            sheet.cell(row=current_row, column=4, value=region_server.gc_time)
            current_row += 1
    if not report_dir.exists():
        report_dir.mkdir(parents=True, exist_ok=True)
    workbook.save(report_dir / f'report-{datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx')
